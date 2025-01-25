from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = "supersecretkey"

# Excel dosyalarını kontrol et ve yoksa oluştur
excel_file = "HayvanBakimKayitlari.xlsx"
ilaclar_file = "Ilaclar.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bakım Kayıtları"
    ws.append([
        "ID", "Hayvan Türü", "Başlama Tarihi", "İlgilenen Kişi",
        "Tedavi Durumu", "Bitiş Tarihi", "İlaçlar", "Notlar"
    ])
    wb.save(excel_file)

if not os.path.exists(ilaclar_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "İlaçlar"
    ws.append(["İlaç Adı"])
    wb.save(ilaclar_file)


# Ana sayfa
@app.route("/")
def index():
    # Bakım kayıtlarını getir
    wb = load_workbook(excel_file)
    ws = wb["Bakım Kayıtları"]
    bakim_kayitlari = [{
        "ID": row[0],
        "Tur": row[1],
        "Baslama": row[2],
        "Kisi": row[3],
        "Durum": row[4],
        "Bitis": row[5],
        "Ilaclar": row[6],
        "Notlar": row[7],
    } for row in ws.iter_rows(min_row=2, values_only=True)]
    return render_template("index.html", bakim_kayitlari=bakim_kayitlari)


# Yeni kayıt ekleme
@app.route("/add", methods=["POST"])
def add_record():
    if request.method == "POST":
        tur = request.form["tur"].capitalize()
        baslama = request.form["baslama"]
        kisi = request.form["kisi"].capitalize()
        durum = request.form["durum"]
        bitis = request.form.get("bitis", "")
        ilaclar = request.form.get("ilaclar", "")
        notlar = request.form.get("notlar", "")

        if not tur or not baslama or not kisi or not durum:
            flash("Lütfen zorunlu alanları doldurun!", "danger")
            return redirect(url_for("index"))

        try:
            datetime.strptime(baslama, "%Y-%m-%d")
        except ValueError:
            flash("Başlama tarihi geçerli bir formatta değil! (YYYY-MM-DD)",
                  "danger")
            return redirect(url_for("index"))

        if bitis:
            try:
                datetime.strptime(bitis, "%Y-%m-%d")
            except ValueError:
                flash("Bitiş tarihi geçerli bir formatta değil! (YYYY-MM-DD)",
                      "danger")
                return redirect(url_for("index"))

        wb = load_workbook(excel_file)
        ws = wb["Bakım Kayıtları"]

        id = 1
        ids = [
            row[0]
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)
            if row[0]
        ]
        if ids:
            id = max(ids) + 1

        ws.append([id, tur, baslama, kisi, durum, bitis, ilaclar, notlar])
        wb.save(excel_file)

        flash("Kayıt başarıyla eklendi!", "success")
        return redirect(url_for("index"))


# Kayıt silme
@app.route("/delete/<int:record_id>")
def delete_record(record_id):
    wb = load_workbook(excel_file)
    ws = wb["Bakım Kayıtları"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == record_id:
            ws.delete_rows(row[0].row)
            wb.save(excel_file)
            flash("Kayıt başarıyla silindi!", "success")
            break
    else:
        flash("Kayıt bulunamadı!", "danger")

    return redirect(url_for("index"))


# İlaçları yönetme
@app.route("/ilaclar")
def ilaclar():
    wb = load_workbook(ilaclar_file)
    ws = wb["İlaçlar"]
    ilac_listesi = [
        row[0] for row in ws.iter_rows(min_row=2, values_only=True)
    ]
    return render_template("ilaclar.html", ilac_listesi=ilac_listesi)


@app.route("/add_ilac", methods=["POST"])
def add_ilac():
    if request.method == "POST":
        ilac_adi = request.form["ilac_adi"].capitalize()
        if not ilac_adi:
            flash("Lütfen ilaç adını girin!", "danger")
            return redirect(url_for("ilaclar"))

        wb = load_workbook(ilaclar_file)
        ws = wb["İlaçlar"]
        ws.append([ilac_adi])
        wb.save(ilaclar_file)

        flash("İlaç başarıyla eklendi!", "success")
        return redirect(url_for("ilaclar"))


@app.route("/delete_ilac/<ilac_adi>")
def delete_ilac(ilac_adi):
    wb = load_workbook(ilaclar_file)
    ws = wb["İlaçlar"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == ilac_adi:
            ws.delete_rows(row[0].row)
            wb.save(ilaclar_file)
            flash("İlaç başarıyla silindi!", "success")
            break
    else:
        flash("İlaç bulunamadı!", "danger")

    return redirect(url_for("ilaclar"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
